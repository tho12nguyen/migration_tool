import os
import difflib
import re
import openpyxl

def read_excel_data(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    excel_data = {}
    for sheet_name in ["SELECT", "INSERT", "UPDATE", "DELETE"]:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]
        data = {}
        for row in sheet.iter_rows(min_row=2):  # B? header
            b_val = row[1].value  # C?t B
            l_val = row[11].value  # C?t L

            if isinstance(b_val, str):
                match = re.search(r'\d+', b_val)
                if match:
                    b_val = int(match.group())

            if isinstance(b_val, int) and b_val >= 1:
                data[b_val] = l_val
        excel_data[sheet_name] = data
    return excel_data

def read_file_lines(file_path):
    encodings = ['cp932', 'shift_jis', 'utf-8', 'euc_jp', 'latin-1']
    for enc in encodings:
        try:
            with open(file_path, 'r', encoding=enc, errors='ignore') as f:
                return f.readlines()
        except Exception:
            continue

    try:
        with open(file_path, 'rb') as f:
            content = f.read()
        return content.decode('cp932', errors='ignore').splitlines(True)
    except Exception:
        print(f"? Kh?ng th? ??c file: {file_path}")
        return []

def find_cast_in_diff(file1_path, file2_path, folder_type, case_number, excel_status):
    lines1 = read_file_lines(file1_path)
    lines2 = read_file_lines(file2_path)
    if not lines1 or not lines2:
        return [], None

    diff = list(difflib.unified_diff(lines1, lines2, fromfile=file1_path, tofile=file2_path, lineterm=''))
    results = []
    result_logs = []
    for line in diff:
        if line.startswith(('+', '-')) and re.search(r'cast', line, re.IGNORECASE):
            clean_line = line[1:].strip()
            results.append(clean_line)

            result_logs.append(f"[{folder_type}] Case {case_number} (Excel: {excel_status}) -> {line[0]} {clean_line}")

    return (results, result_logs)

def extract_case_number(folder_name):
    # H? tr? c? "No.30" v? "No_30"
    match = re.search(r'No[\._-]?\s*0*(\d+)', folder_name, re.IGNORECASE)
    return int(match.group(1)) if match else None

def scan_folder_for_cast(folder_path, excel_data):
    cast_reports = {}
    for root, _, files in os.walk(folder_path):
        folder_type = None
        for key in excel_data.keys():
            if key in root.upper():
                folder_type = key
                break
        case_number = extract_case_number(os.path.basename(root))
        for file in files:
            if '_after' in file.lower():
                continue
            base_path = os.path.join(root, file)
            name, ext = os.path.splitext(file)
            after_file = name + '_after' + ext
            after_path = os.path.join(root, after_file)

            if os.path.exists(after_path):
                status = ""
                if folder_type and case_number:
                    status = excel_data.get(folder_type, {}).get(case_number, "")
                (cast_lines, cast_logs) = find_cast_in_diff(base_path, after_path, folder_type, case_number, status)
                if cast_lines and ord(str(status).strip())== 215:
                    cast_reports[base_path] = {"lines": cast_lines, "status": status}
    return (cast_reports, cast_logs)

